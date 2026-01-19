"""
Export API Routes

Handles exporting chat history to various formats (JSON, CSV, Excel).
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse
from uuid import UUID
from pathlib import Path
import json
import csv
from datetime import datetime
from typing import List, Dict

from models.schemas import ExportRequest, ExportResponse
from database.connection import get_db, DatabaseManager

router = APIRouter(prefix="/api/export", tags=["export"])

# Export directory
EXPORT_DIR = Path(__file__).parent.parent.parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)


# ============================================================================
# EXPORT SESSION
# ============================================================================

@router.post("/session", response_model=ExportResponse)
async def export_session(
    request: ExportRequest,
    db: DatabaseManager = Depends(get_db)
):
    """
    Export chat session to file

    Args:
        request: Export request with format and options
        db: Database manager

    Returns:
        Export response with download URL
    """
    try:
        # Get session data
        async with db.get_cursor() as cursor:
            # Get session info
            cursor.execute(
                """
                SELECT session_id, company_id, company_name, created_at
                FROM chat_sessions
                WHERE session_id = %s
                """,
                (str(request.session_id),)
            )
            session_row = cursor.fetchone()

            if not session_row:
                raise HTTPException(status_code=404, detail="Session not found")

            # Get all messages
            cursor.execute(
                """
                SELECT message_id, role, content, query_metadata, created_at
                FROM chat_messages
                WHERE session_id = %s
                ORDER BY created_at ASC
                """,
                (str(request.session_id),)
            )
            messages = cursor.fetchall()

            # Get feedback if requested
            feedback_map = {}
            if request.include_feedback:
                cursor.execute(
                    """
                    SELECT message_id, feedback_score, feedback_timestamp
                    FROM feedback_responses
                    WHERE session_id = %s
                    """,
                    (str(request.session_id),)
                )
                for row in cursor.fetchall():
                    feedback_map[row['message_id']] = {
                        'score': float(row['feedback_score']),
                        'timestamp': row['feedback_timestamp'].isoformat()
                    }

        # Generate export file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"session_{request.session_id}_{timestamp}.{request.export_format}"
        file_path = EXPORT_DIR / filename

        if request.export_format == "json":
            file_size = _export_to_json(file_path, session_row, messages, feedback_map, request)
        elif request.export_format == "csv":
            file_size = _export_to_csv(file_path, session_row, messages, feedback_map, request)
        elif request.export_format == "xlsx":
            file_size = _export_to_excel(file_path, session_row, messages, feedback_map, request)
        else:
            raise HTTPException(status_code=400, detail="Invalid export format")

        # Save export record
        async with db.get_cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO query_exports (session_id, export_format, file_path, file_size_bytes, message_count)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING export_id, created_at
                """,
                (str(request.session_id), request.export_format, str(file_path), file_size, len(messages))
            )
            export_row = cursor.fetchone()

        return ExportResponse(
            export_id=export_row['export_id'],
            session_id=request.session_id,
            export_format=request.export_format,
            file_path=str(file_path),
            file_size_bytes=file_size,
            message_count=len(messages),
            created_at=export_row['created_at'],
            download_url=f"/api/export/download/{export_row['export_id']}"
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        with open("export_debug_log.txt", "a") as f:
            f.write(f"\n--- Export Error {datetime.now()} ---\n")
            f.write(str(e))
            f.write("\n")
            f.write(traceback.format_exc())
        print(f"Export Error: {str(e)}") # Add logging
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/download/{export_id}")
async def download_export(
    export_id: int,
    db: DatabaseManager = Depends(get_db)
):
    """
    Download exported file

    Args:
        export_id: Export ID
        db: Database manager

    Returns:
        File download response
    """
    try:
        async with db.get_cursor() as cursor:
            cursor.execute(
                """
                SELECT export_id, file_path, export_format
                FROM query_exports
                WHERE export_id = %s
                """,
                (export_id,)
            )
            row = cursor.fetchone()

            if not row:
                raise HTTPException(status_code=404, detail="Export not found")

            file_path = Path(row['file_path'])

            if not file_path.exists():
                raise HTTPException(status_code=404, detail="Export file not found")

            # Update downloaded timestamp
            cursor.execute(
                "UPDATE query_exports SET downloaded_at = NOW() WHERE export_id = %s",
                (export_id,)
            )

            return FileResponse(
                path=file_path,
                filename=file_path.name,
                media_type=_get_media_type(row['export_format'])
            )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")


# ============================================================================
# EXPORT FORMAT HANDLERS
# ============================================================================

def _export_to_json(
    file_path: Path,
    session: Dict,
    messages: List[Dict],
    feedback_map: Dict,
    request: ExportRequest
) -> int:
    """Export to JSON format"""
    export_data = {
        'session': {
            'session_id': str(session['session_id']),
            'company_id': session['company_id'],
            'company_name': session['company_name'],
            'created_at': session['created_at'].isoformat()
        },
        'messages': []
    }

    for msg in messages:
        msg_data = {
            'message_id': msg['message_id'],
            'role': msg['role'],
            'content': msg['content'],
            'created_at': msg['created_at'].isoformat()
        }

        # Add sources if requested
        if request.include_sources and msg['query_metadata']:
            metadata = msg['query_metadata']
            msg_data['sources'] = metadata.get('sources', [])
            msg_data['model_used'] = metadata.get('model_used')
            msg_data['retrieval_tier'] = metadata.get('retrieval_tier_used')

        # Add feedback if available
        if msg['message_id'] in feedback_map:
            msg_data['feedback'] = feedback_map[msg['message_id']]

        export_data['messages'].append(msg_data)

    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False)

    return file_path.stat().st_size


def _export_to_csv(
    file_path: Path,
    session: Dict,
    messages: List[Dict],
    feedback_map: Dict,
    request: ExportRequest
) -> int:
    """Export to CSV format"""
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Header
        headers = ['Message ID', 'Role', 'Content', 'Timestamp']
        if request.include_feedback:
            headers.extend(['Feedback Score', 'Feedback Time'])
        if request.include_sources:
            headers.append('Model Used')

        writer.writerow(headers)

        # Data rows
        for msg in messages:
            row = [
                msg['message_id'],
                msg['role'],
                msg['content'],
                msg['created_at'].isoformat()
            ]

            if request.include_feedback:
                feedback = feedback_map.get(msg['message_id'], {})
                row.extend([
                    feedback.get('score', ''),
                    feedback.get('timestamp', '')
                ])

            if request.include_sources and msg['query_metadata']:
                row.append(msg['query_metadata'].get('model_used', ''))

            writer.writerow(row)

    return file_path.stat().st_size


def _export_to_excel(
    file_path: Path,
    session: Dict,
    messages: List[Dict],
    feedback_map: Dict,
    request: ExportRequest
) -> int:
    """Export to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chat History"

        # Headers
        headers = ['Message ID', 'User Question', 'Assistant Response', 'Result/Answer', 'Timestamp', 'Feedback Score', 'Model Used']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Process messages to pair User -> Assistant
        # We iterate and find 'user' messages, then look for the next 'assistant' message
        
        # Sort messages by ID to ensure order
        sorted_messages = sorted(messages, key=lambda x: x['message_id'])
        
        row_idx = 2
        i = 0
        while i < len(sorted_messages):
            msg = sorted_messages[i]
            
            if msg['role'] == 'user':
                user_msg = msg
                assistant_msg = None
                
                # Look ahead for assistant response
                if i + 1 < len(sorted_messages) and sorted_messages[i+1]['role'] == 'assistant':
                    assistant_msg = sorted_messages[i+1]
                    i += 1 # Skip next message as we consumed it
                
                # Write Row
                # ID
                ws.cell(row=row_idx, column=1, value=user_msg['message_id'])
                
                # User Question
                ws.cell(row=row_idx, column=2, value=user_msg['content'])
                
                if assistant_msg:
                    # Assistant Response (Result)
                    ws.cell(row=row_idx, column=3, value="Assistant") # Just a label or can be empty
                    ws.cell(row=row_idx, column=4, value=assistant_msg['content'])
                    ws.cell(row=row_idx, column=5, value=assistant_msg['created_at'].isoformat())
                    
                    # Feedback
                    if request.include_feedback:
                        feedback = feedback_map.get(assistant_msg['message_id'], {})
                        ws.cell(row=row_idx, column=6, value=feedback.get('score', ''))
                    
                    # Model
                    if request.include_sources and assistant_msg['query_metadata']:
                         ws.cell(row=row_idx, column=7, value=assistant_msg['query_metadata'].get('model_used', ''))
                
                else:
                    # Unanswered user message?
                     ws.cell(row=row_idx, column=5, value=user_msg['created_at'].isoformat())

                row_idx += 1
            
            elif msg['role'] == 'assistant':
                # Orphaned assistant message or system message?
                # For now, print it on its own row if not consumed
                ws.cell(row=row_idx, column=1, value=msg['message_id'])
                ws.cell(row=row_idx, column=3, value="Assistant")
                ws.cell(row=row_idx, column=4, value=msg['content'])
                ws.cell(row=row_idx, column=5, value=msg['created_at'].isoformat())
                row_idx += 1
                
            i += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 80
        ws.column_dimensions['E'].width = 20

        wb.save(file_path)
        return file_path.stat().st_size

    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not installed")


def _get_media_type(export_format: str) -> str:
    """Get media type for export format"""
    media_types = {
        'json': 'application/json',
        'csv': 'text/csv',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    return media_types.get(export_format, 'application/octet-stream')
